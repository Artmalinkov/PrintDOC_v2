'''
Основной функционал приложения по автозамене данных в шаблонах Word
'''
import os
import datetime
import win32api
from docxtpl import DocxTemplate
from openpyxl import load_workbook
import yaml
import pyodbc


def get_config(key):
    '''
    Функция возвращает значение параметра
    :param key: наименование ключа
    :return: value: значение словаря с ключом key
    '''
    with open('config.yaml', 'r', encoding='utf-8') as file:
        tmp_data = yaml.safe_load(file)
    return tmp_data[key]


def set_config(key, new_value):
    '''
    Функция устанавливает новое значение в файле конфигурации
    :param key: значение параметра в файле конфигурации
    :param new_value: новое значение параметра
    :return: None
    '''
    with open('config.yaml', 'r', encoding='utf-8') as file:
        tmp_data = yaml.safe_load(file)
        tmp_data[key] = new_value
    with open('config.yaml', 'w', encoding='utf-8') as file:
        yaml.dump(tmp_data, file)


def get_start(WORKBOOK_PATH: str, NAME_WORKSHEET: str, TABLE_NAME: str, TEST_DOCTPL_PATH: str):
    '''
    Функция проводит первоначальное определение основных объектов с которыми с последующем будем работать
    :param WORKBOOK_PATH:
    :param NAME_WORKSHEET:
    :param TABLE_NAME:
    :param TEST_DOCTPL_PATH:
    :return: wb, ws, table, doc_tpl
    '''

    # Загружаем рабочую книгу
    wb = load_workbook(WORKBOOK_PATH)

    # Определение рабочего листа
    ws = wb[NAME_WORKSHEET]

    # Определяем тяблицу
    table = ws.tables[TABLE_NAME]

    # Определяем шаблон Word
    doc_tpl = DocxTemplate(TEST_DOCTPL_PATH)

    return wb, ws, table, doc_tpl


def get_dict_row(table, row):
    '''
    Функция формирования словаря из значений в строке
    :param table: объект таблицы на рабочем листе
    :param row: объект строки в таблице
    :return: dict_row - сформированный словарь: ключ - заголовок столбца, значение - ячейка
    '''
    dict_row = {}
    for cell in row:
        # Преобразование даты
        if type(cell.value) == datetime.datetime:
            cell.value = cell.value.strftime('%d.%m.%Y')
        # Наполнение словаря
        dict_row[table.column_names[cell.column - 1]] = cell.value
    return dict_row


def iteration_row(wb, ws, table, doc_tpl, MARK):
    '''
    Функция просмотра строк на рабочем листе в таблице.
    :param MARK: маркер по котороку определяется необходимость печати
    :param ws: рабочий лист
    :param table: рабочая таблица
    :return:
    '''

    # Запускаем процедуру просмотра строк в таблице.
    for row in ws[table.ref]:

        # Для каждой строки формируем словарь
        dict_row = get_dict_row(table, row)

        # Если в нужном поле стоит единичка...
        if dict_row[PRINT_COLUMN_NAME] == MARK:

            # Запускаем процедуру замены
            doc_tpl.render(dict_row)

            # Если нужно - сохраняем, потом если нужно - распечатываем.
            if NEED_SAVE == True:
                save_doc_with_name(doc_tpl, dict_row)

                # Если нужно - распечатываем
                if NEED_PRINT == True:
                    print_doc(dict_row)

            # Если нужно - заменяем MARK на текущую дату
            if NEED_CHANGE_NOW_DATE == True:
                change_now_date(wb, table, row, PRINT_COLUMN_NAME, NEED_WB_SAVE)


def change_now_date(wb, table, row, CHANGE_COLUMN_NAME, NEED_WB_SAVE):
    '''
    Функция заменяет значение ячейки в строке на текущую дату.
    :param wb: рабочая книга
    :param table: рабочая таблица
    :param row: исследуемая строка
    :param CHANGE_COLUMN_NAME: наименование столбца в котором нужно заменить значение
    :param NEED_WB_SAVE: определяет, нужно ли сохранять рабочую книгу после манипуляций
    :return:
    '''
    index_change_column = table.column_names.index(CHANGE_COLUMN_NAME)
    row[index_change_column].value = datetime.datetime.now().strftime('%d.%m.%Y')
    if NEED_WB_SAVE == True:
        wb.save(WORKBOOK_PATH)


def save_doc_with_name(changed_doc, dict_row):
    '''
    Функция сохранения измененённого документа
    :param changed_doc: изменённый документ
    :param context: словарь нужной строки
    :return:
    '''
    doc_name = RESULT_DOC_DIR + '\\' + f"{dict_row['Фамилия']}{dict_row['Имя']}.docx"
    changed_doc.save(doc_name)


def print_doc(dict_row):
    '''
    Функция распечатки документа
    :param dict_row: словарь нужной строки
    :return:
    '''

    doc_name = f"done/{dict_row['Фамилия']}{dict_row['Имя']}.docx"

    full_filepath = os.path.abspath(doc_name)
    win32api.ShellExecute(0, 'print', full_filepath, None, '.', 0)


def print_doc(dict_row):
    '''
    Функция распечатки документа
    :param dict_row: словарь нужной строки
    :return:
    '''
    full_filepath = RESULT_DOC_DIR + '\\' + f"{dict_row['Фамилия']}{dict_row['Имя']}.docx"
    win32api.ShellExecute(0, 'print', full_filepath, None, '.', 0)


def connect_to_db():
    '''
    Функция возвращает объект курсора после подключения к базе данных.
    :param filepath_db: Путь к базе данных
    :return: conn: объект подключения
             cursor - объект курсора
    '''
    # Создание подключения к Базе данных
    driver_db = 'Microsoft Access Driver (*.mdb, *.accdb)'
    FILEPATH_DB = r'C:\PythonProjects\Print_AS\attachments\test_db.accdb'
    user_db = ''
    password_db = ''
    connection_string = f'DRIVER={driver_db};DBQ={FILEPATH_DB};UID={user_db};PWD={password_db}'
    conn = pyodbc.connect(connection_string)
    # Cоздание курсора
    cursor = conn.cursor()
    return conn, cursor


def get_headers(cursor, DB_TABLE_NAME):
    '''
    Получение списка заголовков таблицы
    :param table_name: наименование таблицы
    :param cursor: объект курсора
    :return: headers - список заголовков
    '''
    headers = []
    for row in cursor.columns(DB_TABLE_NAME):
        headers.append(row[3])
    return headers


def get_headers_str(cursor, DB_TABLE_NAME):
    '''
    Функция возвращает строку заголовков для последующего формирования запроса SQL на добавление данных
    :param cursor: объект курсора базы данных
    :param DB_TABLE_NAME: наименование таблицы базы данных
    :return: headers_str: текстовая строка заголовков в нужном формате
    '''
    headers = get_headers(cursor, DB_TABLE_NAME)
    headers_str = ''
    for item in headers[1:]:
        headers_str = headers_str + ', [' + item + ']'
    headers_str = headers_str[2:]
    return headers_str


def get_SQL_query(cursor, DB_TABLE_NAME):
    '''
    Функция возвращает переработанную строку SQL-запроса на основании объекта курсора и наименовании таблицы
    :param cursor: объект курсора базы данных
    :param DB_TABLE_NAME: наименование таблицы в базе данных
    :return: SQL_str - строка запроса
    '''
    # Получение списка заголовка в таблице
    headers = get_headers(cursor, DB_TABLE_NAME)

    # Определение количества вопросительных знаков
    symb = ('?,' * (len(headers) - 1))[:-1]

    # Формирование строки заголовков в нужном формате
    headers_str = get_headers_str(cursor, DB_TABLE_NAME)

    # Формирование итоговой строки SQL-запроса
    SQL_str = f'INSERT INTO {DB_TABLE_NAME} ({headers_str}) VALUES ({symb})'

    return SQL_str


def print_table_db(cursor, DB_TABLE_NAME):
    '''
    Функция выводит в терминал таблицу базы данных
    :param cursor: объект курсора базы данных
    :param DB_TABLE_NAME: наименование таблицы базы данных
    :return:
    '''

    # Выборка для проверки результата
    cursor.execute(f'SELECT * FROM {DB_TABLE_NAME}')
    # Вывести на экран результаты
    for row in cursor.fetchall():
        print(row)


def insert_row_to_db(cursor, DB_TABLE_NAME, row):
    '''
    Функция вставки строки в базу данных
    :param cursor: объект курсора
    :param DB_TABLE_NAME: наименование таблицы
    :param row: строка данных
    :return:
    '''
    SQL_str = get_SQL_query(cursor, DB_TABLE_NAME)
    cursor.execute(SQL_str, row)
    conn.commit()


def main():
    wb, ws, table, doc_tpl = get_start(WORKBOOK_PATH, NAME_WORKSHEET, TABLE_NAME, DOCTPL_PATH)
    iteration_row(wb, ws, table, doc_tpl, MARK)


# Путь к рабочему файлу Excel
WORKBOOK_PATH = get_config('WORKBOOK_PATH')

# Наименование рабочего листа
NAME_WORKSHEET = get_config('NAME_WORKSHEET')

# Наименование таблицы на рабочем листе
TABLE_NAME = get_config('TABLE_NAME')

# Столбец по которому определяется необходимость печати документа
PRINT_COLUMN_NAME = get_config('PRINT_COLUMN_NAME')

# Маркер по отслеживанию необходимости печати
MARK = get_config('MARK')

# Путь к рабочему шаблону Word
DOCTPL_PATH = get_config('DOCTPL_PATH')

# Путь к папке, в которой будут лежать результаты
RESULT_DOC_DIR = get_config('RESULT_DOC_DIR')

# Необходимость выполнения распечатывания документов
NEED_PRINT = get_config('NEED_PRINT')

# Необходимость сохранения изменённого документа Word в файл
NEED_SAVE = get_config('NEED_SAVE')

# Необходимость замены маркера на сегодняшнюю дату
NEED_CHANGE_NOW_DATE = get_config('NEED_CHANGE_NOW_DATE')

# Необходимость сохранения рабочей книги после манипуляций
NEED_WB_SAVE = get_config('NEED_WB_SAVE')

# Наименование таблицы в базе данных
DB_TABLE_NAME = get_config('DB_TABLE_NAME')

# Путь до базы данных
# FILEPATH_DB = get_config('FILEPATH_DB')

if __name__ == '__main__':
    main()

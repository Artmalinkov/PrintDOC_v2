from openpyxl import Workbook, load_workbook
import os
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Series, Reference

wb = load_workbook('src_learn/sample.xlsx')
ws = wb['Лист1']
TreeData = [
    ["Тип", "Цвет листа", "Высота"],
    ["Клен", "Красный", 549],
    [" Дуб", "Зеленый", 783],
    ["Сосна", "Зеленый", 1204]]

os.system('start excel.exe "C:\PythonProjects\Print_AS\src_learn\sample.xlsx"')
os.system('start excel.exe "src_learn\sample.xlsx"')


ws['A1'] = 42
for i in range(1,50):
    for j in range(1,50):
        ws.cell(i,j).value = 'Ну привет'

cell_range = ws['A1':'B10']
for row in cell_range:
    for cell in row:
        cell.value = '0'


for cols in ws.iter_cols(3,5,6,9):
    for cell in cols:
        cell.value = 'И'

wb.save("src_learn/sample.xlsx")
os.system('start excel.exe "C:\PythonProjects\Print_AS\src_learn\sample.xlsx"')


# TODO посмотреть руководство по openpyxl

for cell in ws.values:
    print(cell)


# Простое использование
wb = load_workbook('src_learn/sample.xlsx')
ws = wb['Лист1']
TreeData = [
    ["Тип", "Цвет листа", "Высота"],
    ["Клен", "Красный", 549],
    [" Дуб", "Зеленый", 783],
    ["Сосна", "Зеленый", 1204]]
for row in TreeData:
    ws.append(row)
# Сделать шрифт жирным
ft = Font(bold=True)
for row in ws["A1:C1"]:
    for cell in row:
        cell.font = ft

# Создание диаграмм
chart = BarChart()
chart.type = "col"
chart.title = "Tree Height"
chart.y_axis.title='Height (cm)'
chart.x_axis.title = 'Tree Type'
chart.legend = None
data = Reference(ws, min_col=3, min_row=2, max_row=4, max_col=3)
categories = Reference(ws, min_col=1, min_row=2, max_row=4, max_col=1)
chart.add_data(data)
chart.set_categories(categories)
ws.add_chart(chart, "E1")
wb.save("src_learn/TreeData.xlsx")

# Работа со стилями
# TODO При необходимости вернуться и пересмотреть


# Работа с форматированным текстом
# TODO При необходимости вернуться и пересмотреть

# Условное форматирование
# TODO При необходимости вернуться и пересмотреть

# Вставка и удаление строк и столбцов, перемещение диапазонов ячеек
ws.insert_rows(2)
# TODO При необходимости вернуться и пересмотреть

# Дополнительные свойства листа
# TODO При необходимости вернуться и пересмотреть

# Проверка ячеек
# TODO При необходимости вернуться и пересмотреть

# Таблицы рабочих листов
# Создание таблицы
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = load_workbook('src_learn/sample.xlsx')
wb.iso_dates = True
ws = wb['Лист1']
table = ws.tables['Таблица1']
# table.ref - Диапазон таблицы
# Перебор значений в таблице
for row in ws[table.ref]:
    for cell in row:
        print(cell.value, end='|')
    print()
# TODO Разобраться с форматом даты

wb.save("src_learn/sample.xlsx")
os.system('start excel.exe "src_learn\sample.xlsx"')



